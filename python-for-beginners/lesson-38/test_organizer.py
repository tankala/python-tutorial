import openpyxl
from Question import Question
from random import randint

def get_questions_and_answers():
    questions = []
    workbook_obj = openpyxl.load_workbook("mcq.xlsx")
    sheet_obj = workbook_obj.active
    max_row = sheet_obj.max_row

    for row_number in range(2, max_row+1):
        question = sheet_obj.cell(row = row_number, column = 1).value
        answer = sheet_obj.cell(row = row_number, column = 2).value
        current_question_obj = Question(question, answer)
        questions.append(current_question_obj)
    
    return questions

def start_test_and_give_score():
    questions = get_questions_and_answers()
    score = 0
    for turn in range(3):
        random_question_index = randint(0, len(questions)-1)
        selected_question = questions[random_question_index]
        user_answer = input(selected_question.question + "\n\n: ")
        if user_answer == selected_question.answer:
            score += 1
    return score
